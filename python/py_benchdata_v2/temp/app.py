import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.scrolledtext as st
import tkinter.messagebox as mb
import tkinter.font as tkFont
import win32com.client as client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import datetime
import subprocess
import json
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from configparser import ConfigParser
from icecream import ic
from inspect import currentframe
import sys
import subprocess

plt.rcParams["axes.prop_cycle"] = plt.cycler(color=["#00FF00", "#FF0000", "#0000FF", "#FFFF00", "#FF00FF"])

New_Report_file = ''
report_file = ''
report_date = datetime.datetime.now()
bench_count = 0
upcoming_bench_count = 0
available_within_15_days = 0
available_in_16_to_30_days = 0
available_in_31_to_60_days = 0
available_in_61_to_90_days = 0
available_after_90_days = 0
supply_file =''
Confirmed = 0
Proposed = 0 
Available = 0
sub_total = 0
sender_name = ""
notes = ""
plt_key = '2024-09-01 12:34:56'
tower_list = ['']
selected_tower = ""
plt1_data = { 
                "CONFIRMED": 0,
                "PROPOSED": 0,
                "AVAILABLE": 0,
            }
plt2_data = {
                "WITHIN15": 0,
                "WITHIN30": 0,
                "WITHIN60": 0,
                "WITHIN90": 0,
                "AFTER90": 0   
            }


# ic.enable()
ic.disable()

# -----------

def read_conf():
    try:
        conf = ConfigParser()
        if (os.path.isfile('config.ini')):
            conf.read('config.ini')
            supply_sheet = conf.get('Default', 'supply_sheet')
            report_sheet = conf.get('Default', 'report_sheet')
            demand_sheet = conf.get('Default', 'demand_sheet')
            utilization_sheet = conf.get('Default', 'utilization_sheet')
            skill_sheet = conf.get('Default', 'skill_sheet')
            sender_name = conf.get('Default', "sender_name")
            notes = conf.get("Default", "notes")
            supply_sheet = os.path.join(os.getcwd(), supply_sheet)
            report_sheet = os.path.join(os.getcwd(), report_sheet)
            demand_sheet = os.path.join(os.getcwd(), demand_sheet)
            utilization_sheet = os.path.join(os.getcwd(), utilization_sheet)
            skill_sheet = os.path.join(os.getcwd(), skill_sheet)
            

            return supply_sheet, report_sheet, demand_sheet, utilization_sheet, skill_sheet, sender_name, notes
        else:
            print ("\nConfig.ini file not found in current directoy.")
            return None    
    except Exception as e:
        mb.showerror("Error", e)
        return None
    
# ---------------
def get_line():
    # Ln = currentframe().f_back.f_lineno
    return sys._getframe().f_back.f_lineno
    
# -------------
ic.configureOutput(prefix=f'Debug:{get_line()} | ')
# ------------------------
def get_plt_data(key):
    bench = { 
                "CONFIRMED": 0,
                "PROPOSED": 0,
                "AVAILABLE": 0,
            }
    upcoming = {
                "WITHIN15": 0,
                "WITHIN30": 0,
                "WITHIN60": 0,
                "WITHIN90": 0,
                "AFTER90": 0
            }
    try:
        db= read_history(key=key)
    except Exception as e:
        mb.INFO("No historical data found.")
        return bench, upcoming

    bench["CONFIRMED"] = db["CONFIRMED"]
    bench["PROPOSED"] = db["PROPOSED"]
    bench["AVAILABLE"] = db["AVAILABLE"]

    upcoming["WITHIN15"] = db["WITHIN15"]  
    upcoming["WITHIN30"] = db["WITHIN30"]
    upcoming["WITHIN60"] = db["WITHIN60"]
    upcoming["WITHIN90"] = db["WITHIN90"]
    upcoming["AFTER90"] = db["AFTER90"]

    return bench, upcoming
# --------------------------
# -------------------------

# ----------------------

def open_excel_file(New_Report_file):
    file_to_open= os.getcwd() + "\\" +New_Report_file
    try:
            
        if os.path.exists(file_to_open):
            subprocess.Popen(["start", "excel", file_to_open], shell=True)
        else:
            raise FileNotFoundError("Excel file not found.")
    except Exception as e:
        mb.showerror("Error", e)
        return None
# ------------------------
def read_body(body_path):
    try:
        with open(body_path, 'r', encoding='utf-8') as file:
            content = file.read()
            return content
    except FileNotFoundError:
        mb.showerror("Error", "File not found.")
        return None
    except Exception as e:
        mb.showerror("Error", e)
        return None
# ------------------------
def save_history(key, json_data):
    history_path = os.path.join(os.getcwd(), "history.json")
    if not os.path.exists(history_path):
        mb.showerror("Error", "History database not found.")
        return None
    
    try:
        with open('history.json', 'r') as file:
            historical_data = json.load(file)
    except FileNotFoundError:
        historical_data = {}

    historical_data[key] = json_data
    try:
        with open('history.json', 'w') as file:
            json.dump(historical_data, file, indent=4)
    except Exception as e:
        mb.showerror("Error", e)
        return None 
    
def read_history(key):
    history_path = os.path.join(os.getcwd(), "history.json")
    if not os.path.exists(history_path):
        mb.showerror("Error", "History database not found.")
        return None
    
    try:
        with open('history.json', 'r') as file:
            historical_data = json.load(file)
    except FileNotFoundError:
        historical_data = {}

    if key in historical_data:
        return historical_data[key]
    else:
        return None
        
# ------------------------
def read_email_notes():
    conf = ConfigParser()
    try:
        if (os.path.isfile('config.ini')):
            conf.read('config.ini')
            sender_name = conf.get('Default', "sender_name")
            notes = conf.get("Default", "notes")
            return sender_name, notes
        else:
            raise Exception("Config read issue.")    
    except Exception as e:
        mb.showerror(e)
        return None   

def SendEmailbutton():
    global supply_sheet
    global bench_count
    global upcoming_bench_count
    global available_within_15_days
    global available_in_31_to_60_days
    global available_in_61_to_90_days
    global available_after_90_days
    
    if New_Report_file  == '':
        mb.showerror("Error", "No data has been processed today.\nSending a blank email may not be the best choice.")
        return None
    sender_name, notes = read_email_notes()
    excel_path= E1.get()
    html_json_path =os.path.join(os.getcwd(), "body.htm")
    email_body = read_body(html_json_path)
    if not email_body:
        mb.showerror("Error", "Email body not found.")  
        return
    
    now = datetime.datetime.now()
    email_date = now.strftime("%Y-%d-%m %H:%M:%S")

    variables_list = { "REPORT_DATE": "09-01-2024 23:17",
                      "TOWER":"",
                      "SENDER": "",
                      "NOTES": "",
                      "SUPPLY_FILE": "Jayram_data.xlsx",
                      "B_COUNT": "",
                      "CONFIRMED": "",
                      "PROPOSED": "",
                      "AVAILABLE": "",
                      "UP_COUNT": "",
                      "WITHIN15": "",
                      "WITHIN30": "",
                      "WITHIN60": "",
                      "WITHIN90": "",
                      "AFTER90": "",
                      "SUB_TOTAL": "",
                      "TOTAL": ""
    }

    variables_list["REPORT_DATE"] = email_date
    variables_list["TOWER"] = selected_tower
    variables_list["SENDER"] = sender_name
    variables_list["NOTES"] = notes
    variables_list["SUPPLY_FILE"] = os.path.basename(supply_sheet)
    variables_list["B_COUNT"] = str(bench_count)
    variables_list["CONFIRMED"] = str(Confirmed)
    variables_list["PROPOSED"] = str(Proposed)
    variables_list["AVAILABLE"] = str(Available)
    variables_list["UP_COUNT"] = str(upcoming_bench_count)
    variables_list["WITHIN15"] = str(available_within_15_days)
    variables_list["WITHIN30"] = str(available_in_16_to_30_days)
    variables_list["WITHIN60"] = str(available_in_31_to_60_days)
    variables_list["WITHIN90"] = str(available_in_61_to_90_days)
    variables_list["AFTER90"] = str(available_after_90_days)
    variables_list["SUB_TOTAL"] = str(sub_total)
    variables_list["TOTAL"] = str(bench_count+upcoming_bench_count)

    for variable, value in variables_list.items():
        placeholder = f'%{variable}%'
        email_body = email_body.replace(placeholder, value)

    recipients = read_recipient(excel_path)
    attachment_path =os.path.join(os.getcwd(), New_Report_file)
    subject = f"{selected_tower} Bench data report for {email_date} with attachment for today's review."
    try:
        send_email(subject, email_body, attachment_path, recipients)
        mb.showinfo("Information", "Email Sent Successully.")
        scrolled_text.insert(tk.END, "Email Sent Successully.")
        scrolled_text.see(tk.END)
        scrolled_text.update()
        key = email_date
        save_history(key, variables_list )
    
    except Exception as e:
        mb.showerror("Error", e)
        return None

# ------------ --   
def get_tower_list(supply_sheet):
    try:
        
        xls= pd.ExcelFile(supply_sheet)
        if "Supply" in xls.sheet_names:
            # pd = xls.parse("Supply")
            columns_to_save = ['Skill', 'Serial', 'Practitioner Notes ID', 'JRSS-Primary', 'Roll Off Status', 'BandAsPerSR', 'Bench Aging', 'PAL Dept', 'Reporting GDMIS - PAL Dept','Customer Name','Project Name'] 
            df = pd.read_excel(supply_sheet, sheet_name="Supply" , usecols= columns_to_save, header=0)
            Towers_list = df['Skill'].unique()
            options  = sorted(Towers_list)
            combobox_tower['values'] = options
            combobox_tower.set(options[0])
            combobox_tower.config(width=17)
            # combobox_tower.pack()

            # return options
        else:
            mb.showerror("Error", "Supply sheet not found.")
            return None
    except Exception as e:
        mb.showerror("Error", e)
        return None  
# --------------
def on_tower_select(value):
    global selected_tower
    selected_tower = value
# ------------
def browse_supply():

    file1 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select Supply Sheet :",
        filetypes=(("Excel files", "*.xlsx"))
    )
    if file1:
       supply_sheet_ttk.delete(0, tk.END)
       supply_sheet_ttk.insert(0, file1)
# ------------ --   
def browse_tracker():
    file2 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select Report Sheet :",
        filetypes=(("Excel files", "*.xlsx"))
    )
    if file2:
        report_sheet_ttk.delete(0, tk.END)
        report_sheet_ttk.insert(0, file2)            
# --------------
def browse_dm():
    file3 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select Demand Sheet :",
        filetypes=(("Excel files", "*.xlsx"))
    )
    if file3:
        demand_sheet_ttk.delete(0, tk.END)
        demand_sheet_ttk.insert(0, file3)           
# ---------------------
def browse_util():
    file4 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select Utilization Sheet :",
        filetypes=(("Excel files", "*.xlsx"))
    )
    if file4:
        utilization_sheet_ttk.delete(0, tk.END)
        utilization_sheet_ttk.insert(0, file4)   
# -----------------   
def browse_skill():
    file5 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select Skill Sheet :",
        filetypes=(("Excel files", "*.xlsx"))
    )
    if file5:
        skill_sheet_ttk.delete(0, tk.END)
        skill_sheet_ttk.insert(0, file5)   
# -----------------           
def browse_PDL(E1):
    pdl_file = filedialog.askopenfilename( 
        initialdir=os.getcwd(),
        title="Select a file",
        filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*"))
    )
    if pdl_file:
        E1.delete(0, tk.END)
        E1.insert(0, pdl_file)            
# ---------------
def dashboard():
    
    # subprocess.run("streamlit run ./dist/app.py --server.port 8080", shell=True)
    process = subprocess.Popen("streamlit run 10125113.py --server.port 8080", shell=True)

    # process = subprocess.Popen(
    #     "streamlit run ./dist/app.py --server.port 8080",
    #     stdout=subprocess.PIPE,
    #     stderr=subprocess.PIPE,
    #     shell=True
    # )

# ----------------
def update_dashboard(dashboard_file):
    # write file on disks
    dash_file = os.path.join(os.getcwd(),dashboard_file)
    try:
        with open("web.ini", 'w') as f:
            f.write(dash_file)
            f.close()
    except Exception as e:
        return None
        
def GetData():
    global supply_file
    global supply_sheet, report_sheet, demand_sheet, utilization_sheet
    global supply_sheet_base, report_sheet_base, demand_sheet_base, utilization_sheet_base
    global plt1_data, plt2_data

    supply_sheet = supply_sheet_ttk.get()
    report_sheet = report_sheet_ttk.get()
    demand_sheet = demand_sheet_ttk.get()
    utilization_sheet = utilization_sheet_ttk.get()
    skill_sheet = skill_sheet_ttk.get()
    
    supply_file_base = os.path.basename(supply_sheet)
    report_sheet_base = os.path.basename(report_sheet)
    demand_sheet_base = os.path.basename(demand_sheet)
    utilization_sheet_base = os.path.basename(utilization_sheet)
    skill_sheet_base = os.path.basename(skill_sheet)

    logs_str = f"Processing...\nSupply Sheet: {supply_file_base} \nReport Sheet: {report_sheet_base} \nDemand Sheet : {demand_sheet_base} \nUtilization Sheet: {utilization_sheet_base} \nSkills Sheet: {skill_sheet_base} "
    scrolled_text.insert(tk.END, logs_str)
    
    dashboard_file = GetBenchData(supply_sheet, report_sheet, demand_sheet, utilization_sheet, skill_sheet)
    update_dashboard(dashboard_file)

    # self.hyperlink_enabled = True
    if os.path.exists(New_Report_file):
        # hyperlink_font = tkFont.Font(family='Helvetica', size=12, slant='italic', underline=True)
        hyperlink_label.config(text=New_Report_file)
        hyperlink_label.bind("<Button-1>", lambda event: open_excel_file(New_Report_file))
        hyperlink_label.config(fg="white") 
    else:
        hyperlink_label.config(text=New_Report_file)
        hyperlink_label.unbind("<Button-1>")
        hyperlink_label.config(fg="gray")  

# -----------------
def ExcelFormat(file_path, sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    header_fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    blue_border = Border(
        left=Side(style='thin', color='0000FF'),
        right=Side(style='thin', color='0000FF'),
        top=Side(style='thin', color='0000FF'),
        bottom=Side(style='thin', color='0000FF')
    )

    align_left_center = Alignment(horizontal='left', vertical='center')

    for cell in sheet[1]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_left_center
        if cell.column <= 12:
            cell.fill = header_fill_gray
        else:
            cell.fill = header_fill    
    
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18  
        for cell in row:
            cell.border = blue_border
            cell.alignment = align_left_center

    for column in sheet.iter_cols():
        # sheet.column_dimensions[column[0].column_letter].width = 15  
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        if max_length > 50:
            max_length = 50
        if max_length < 5:
            max_length = 5

        adjusted_width = (max_length + 2)  
        sheet.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
# ------------------------
def GetBenchData(supply_sheet, report_sheet, demand_sheet, utilization_sheet, skill_sheet):
    global New_Report_file
    global report_date
    global bench_count
    global upcoming_bench_count
    global available_within_15_days
    global available_in_16_to_30_days
    global available_in_31_to_60_days
    global available_in_61_to_90_days
    global available_after_90_days
    global Confirmed
    global Proposed
    global Available
    global sub_total
    global selected_tower
    global plt1_data
    global plt2_data

    pd.options.mode.chained_assignment = None
    now = datetime.datetime.now()
    report_date = now.strftime("%Y-%m-%d")
    New_Report_file = now.strftime(f"Report_{selected_tower}_%Y%m%d_%H%M%S.xlsx")

    New_sheet_bench = "Bench"
    New_sheet_upcoming_bench = "Upcoming_Bench"
    
    ans = mb.askquestion("Confirmation",f"You are about to process {selected_tower} Tower data?")
    if ans == 'no':
        return None
    # -----------------
    try:
        col_to_read = ['Skill','Serial', 'Name', 'JRSS-Primary', 'Roll Off Status', 'BandAsPerSR', 'Bench Aging', 'PAL Dept', 'Reporting GDMIS - PAL Dept','Customer Name','Project Name'] 
        supply_df = pd.read_excel(io=supply_sheet,engine="openpyxl", sheet_name='Supply', usecols=col_to_read)
        upcoming_df = supply_df.copy()

        ic(supply_df.shape)
    except Exception as e:
        ic(e)
        mb.showerror("Error-no1", e)
        return None 
  
    try:
        missing_values = supply_df['Roll Off Status'].isna().any()
        if missing_values:
            raise Exception("Missing values in 'Roll Off Status' column")
        bench_df_selected = supply_df[(supply_df['Skill'] == selected_tower) & (supply_df['Roll Off Status'] == 'BENCH')].reset_index(drop=True)
        bench_df_selected = bench_df_selected.rename(columns={'Skill': 'Tower'})
        ic(bench_df_selected.shape)
        ic(bench_df_selected.columns)

    except Exception as e:
        ic(e)
        mb.showerror("Error-no2", e)
        return None

    new_columns = ['Serial','Lead', 'Action Owner', 'Status2', 'Account Proposed','Proposed / Confirmed Cased', 'Remarks', 'Comparision']
    
    
    # -- Old data
    try:
        old_tracker_df = pd.read_excel(io=report_sheet,engine="openpyxl", sheet_name= New_sheet_bench, usecols=new_columns)
        bench_df_selected = bench_df_selected.merge(old_tracker_df, on='Serial', how='left')
    except Exception as e: 
        ic(e)
        mb.showerror("Error-no3", e)
        return None
    
    # add utilization
 
    try:
        
        usecols1=['Emp_No', '13 Wk Ute']
        df_uti = pd.read_excel(io=utilization_sheet, sheet_name='Team List & Forecast', header=0 )
        df_uti = df_uti[usecols1]
        df_uti.drop([0, 1], inplace=True)
        df_uti = df_uti.drop_duplicates(subset=['Emp_No'])
        df_uti = df_uti.rename(columns={'Emp_No': 'Serial'})
        df_uti['13 Wk Ute'] = pd.to_numeric(df_uti['13 Wk Ute'], errors='coerce')

        df_uti['13 Wk Ute'] = (df_uti['13 Wk Ute'] * 100).round().astype(int).astype(str) + '%'
        
        bench_df_selected = bench_df_selected.merge(df_uti, on='Serial', how='left')

    except Exception as e:
        ic(e)
        mb.showerror("Error-no4", e)
        return None    
    

# ----------skill--
    try:
        df_skill = pd.read_excel(io=skill_sheet, sheet_name='Skills', usecols=['Talent ID','Skills','Secondary Skills'])
        df_skill = df_skill.drop_duplicates(subset=['Talent ID'])
        df_skill = df_skill.rename(columns={'Talent ID': 'Serial'})
        bench_df_selected = bench_df_selected.merge(df_skill, on='Serial', how='left')
    except Exception as e:
        ic(e)
        mb.showerror("Error-no5", e)
        return None    
    #  NO unique column 
    # try:
    #     header_row =2
    #     take_col = ['Opp/Client Name','Skills / Role','Demand Status','Resource Names']
    #     df_dm = pd.read_excel(io=demand_sheet, sheet_name='DemandTracker', usecols=take_col, header=header_row-1)
    #     df_dm = df_dm.rename(columns={'Resource Names': 'Name'})
    #     df_dm = df_dm.drop_duplicates(subset=['Name'])  
    #     ic(df_dm.shape)
    #     ic(df_dm.columns)
    #     bench_df_selected = bench_df_selected.merge(df_dm, on='Name', how='left')
    #     ic(bench_df_selected.columns)

    # except Exception as e:
    #     ic(e)
    #     mb.showerror("Error-no5", e)
    #     return None    

    bench_df_selected.insert(0, 'SN', range(1, 1 + len(bench_df_selected)))  # add to end

    #  save-
    col_to_save1 = ['SN','Tower','Serial','Name','JRSS-Primary','Roll Off Status','BandAsPerSR','Bench Aging','PAL Dept','Reporting GDMIS - PAL Dept','Customer Name',
                    'Project Name','Lead','Action Owner','Status2','Proposed / Confirmed Cased','Remarks','Comparision', '13 Wk Ute', 'Skills','Secondary Skills']
    bench_df_selected = bench_df_selected[col_to_save1]
    try:
        bench_df_selected.to_excel(New_Report_file, sheet_name= New_sheet_bench, index=False)
        ic(bench_df_selected.shape)
    except Exception as e:
        ic(e)
        mb.showerror("Error-no6", e)
        return None
    
    bench_count = len(bench_df_selected)

    logs_str = f"\nNew Excel Report : {New_Report_file} \n\nTotal Bench Count: {bench_count} "
    scrolled_text.insert(tk.END, logs_str)
   
    # Upcoming Bench # -------
    try:
        missing_values = upcoming_df['Roll Off Status'].isna().any()
        if missing_values:
            raise Exception("Missing values in 'Roll Off Status' column")
        upcoming_bench_df = upcoming_df[(upcoming_df['Skill'] == selected_tower) & (supply_df['Roll Off Status'].str.contains('AVAILABLE'))].reset_index(drop=True)
        upcoming_bench_df = upcoming_bench_df.rename(columns={'Skill': 'Tower'})
        # upcoming_bench_df = upcoming_df[upcoming_df['Roll Off Status'].str.contains('AVAILABLE')]
        
        Upcoming_columns = ['Tower','Serial', 'Name', 'JRSS-Primary', 'Roll Off Status', 'BandAsPerSR', 'Bench Aging', 'PAL Dept', 'Reporting GDMIS - PAL Dept','Customer Name','Project Name'] 
        upcoming_bench_selected = upcoming_bench_df[Upcoming_columns].reset_index(drop=True)

    except Exception as e:
        ic(e)
        mb.showerror("Error-no7", e)
        return None
    
    # add data
    try:
        upcoming_col_list = ['Serial','Lead','Action Owner','Status2','Proposed / Confirmed Cased','Remarks','Comparision']
        old_upcoming_data = pd.read_excel(io=report_sheet,engine="openpyxl", sheet_name= New_sheet_upcoming_bench, usecols=upcoming_col_list)

        new_mergered_df = upcoming_bench_selected.merge(old_upcoming_data, on='Serial', how='left')
        new_mergered_df.insert(0, 'SN', range(1, 1 + len(new_mergered_df)))  # add to end
        
    except Exception as e:
        ic(e)
        mb.showerror("Error-no8", e)
        return None
    # uti
    try:
        new_mergered_df = new_mergered_df.merge(df_uti, on='Serial', how='left')
        new_mergered_df = new_mergered_df.merge(df_skill, on='Serial', how='left')
        col_to_save = ['SN','Tower','Serial','Name','JRSS-Primary','Roll Off Status','BandAsPerSR','Bench Aging','PAL Dept','Reporting GDMIS - PAL Dept','Customer Name','Project Name','Lead','Action Owner','Status2','Proposed / Confirmed Cased','Remarks','Comparision','13 Wk Ute','Skills','Secondary Skills']
        new_mergered_df = new_mergered_df[col_to_save]
    except Exception as e:
        ic(e)    
        mb.showerror("Error-no8", e)
        return None
    try:
        with pd.ExcelWriter(New_Report_file, engine='openpyxl', mode='a') as writer:
            new_mergered_df.to_excel(writer, sheet_name=New_sheet_upcoming_bench, index=False)
    except Exception as e:
        ic(e)
        mb.showerror("Error-no8", e)
        return None

    upcoming_bench_count = len(new_mergered_df)
    logs_str = f"\nUpcoming Bench Count: {upcoming_bench_count}\n"
    # app_instance.scrolled_text.insert(tk.END, logs_str)
    
    df_report = new_mergered_df.copy()
    
    mask_15_days = df_report['Roll Off Status'] == 'AVAILABLE WITHIN 15 DAYS'
    mast_16_30_days = df_report['Roll Off Status'] == 'AVAILABLE IN 16 TO 30 DAYS'
    mask_31_60_days = df_report['Roll Off Status'] == 'AVAILABLE IN 31 TO 60 DAYS'
    mask_61_90_days = df_report['Roll Off Status'] == 'AVAILABLE IN 61 TO 90 DAYS'
    mask_after_90_days = df_report['Roll Off Status'] == 'AVAILABLE AFTER 90 DAYS'

    # Applying the masks to the DataFrame and getting the counts
    available_within_15_days = df_report[mask_15_days].shape[0]
    available_in_16_to_30_days = df_report[mast_16_30_days].shape[0]
    available_in_31_to_60_days = df_report[mask_31_60_days].shape[0]
    available_in_61_to_90_days = df_report[mask_61_90_days].shape[0]
    available_after_90_days = df_report[mask_after_90_days].shape[0]

    sub_total = available_within_15_days + available_in_16_to_30_days + available_in_31_to_60_days + available_in_61_to_90_days

    str_dashboard = f"{selected_tower} Report Date: {report_date}"
    report_frame.config(text=str_dashboard)
    bench_count_value.config(text=str(bench_count))
    upcoming_bench_count_value.config(text=str(upcoming_bench_count))
    total_count.config(text=str(bench_count+upcoming_bench_count))

    vAVAILABLE_WITHIN_15_DAYS.config(text=str(available_within_15_days))
    vAVAILABLE_IN_16_TO_30_DAYS.config(text=str(available_in_16_to_30_days))
    vAVAILABLE_IN_31_TO_60_DAYS.config(text=str(available_in_31_to_60_days))
    vAVAILABLE_IN_61_TO_90_DAYS.config(text=str(available_in_61_to_90_days))
    vAVAILABLE_AFTER_90_DAYS.config(text=str(available_after_90_days))
    logs_str = f"\nReport Generated\n-----------------\n"
    scrolled_text.insert(tk.END, logs_str)
    
    # ------------additional field------
    

    df_bench_sub = pd.read_excel(io=New_Report_file, engine="openpyxl", sheet_name=New_sheet_bench) 
    
    status2_filter = ['Confirmed', 'confirm']
    mask_df_confirmed = df_bench_sub[df_bench_sub['Status2'].isin(status2_filter) | df_bench_sub['Status2'].isna()]
    Confirmed = len(mask_df_confirmed)

    # mask_df_confirmed = df_bench_sub['Status2'].str.lower().isin(['confirmed', 'confirm', ''])
    # mask_df_confirmed = df_bench_sub['Status2'] == 'confirm'
    mask_df_proposed = df_bench_sub['Status2'] == 'Proposed'
    mask_df_available = df_bench_sub['Status2'] == 'Bench'

    # Confirmed = df_bench_sub[mask_df_confirmed].shape[0]
    Proposed = df_bench_sub[mask_df_proposed].shape[0]
    Available = df_bench_sub[mask_df_available].shape[0]
    
    plt1_data["CONFIRMED"] = Confirmed
    plt1_data["PROPOSED"] = Proposed
    plt1_data["AVAILABLE"] = Available

    plt2_data["WITHIN15"] = available_within_15_days
    plt2_data["WITHIN30"] = available_in_16_to_30_days
    plt2_data["WITHIN60"] = available_in_31_to_60_days
    plt2_data["WITHIN90"] = available_in_61_to_90_days
    plt2_data["AFTER90"] = available_after_90_days
    #  merge previoud data
    
    ExcelFormat(New_Report_file, New_sheet_bench)
    ExcelFormat(New_Report_file, New_sheet_upcoming_bench)
    logs_str = f"\nExcel Formatted\n----------------\n"
    scrolled_text.insert(tk.END, logs_str)
    
    ic(plt1_data)
    ic(plt2_data)
    
    if all(value == 0 for value in plt1_data.values()):
        mb.showerror("Error", "Nothing to draw!")
        return None
    else:    
        ax1.clear()
        ax1.pie(list(plt1_data.values()), labels=list(plt1_data.keys()), autopct='%1.1f%%', textprops={'fontsize': 8, 'color': 'white'})
        canvas1.draw()
    if all(value == 0 for value in plt2_data.values()):
        mb.showerror("Error", "Nothing to draw!")
        return None
    else:
        ax2.clear()
        ax2.pie(list(plt2_data.values()), labels=list(plt2_data.keys()), autopct='%1.1f%%', textprops={'fontsize': 8, 'color': 'white'})
        canvas2.draw()
    return New_Report_file
# ------------------------
def send_email(subject, body, attachment_path, recipients):
    outlook = client.Dispatch('Outlook.Application')
    message = outlook.CreateItem(0) 
    message.Subject = subject
    
    message.HTMLBody = body
    
    if attachment_path and os.path.exists(attachment_path):
        message.Attachments.Add(Source=attachment_path)

    
    for recipient in recipients:
        message.Recipients.Add(recipient)

    message.Send()
    # print("Email sent to:", recipients)

def read_recipient(excel_path):
    try:
        df = pd.read_excel(io=excel_path,engine="openpyxl")
        receipient_count = df['Email'].count()
        if receipient_count == 0:
            raise ValueError("No recipients found in the Excel file.")  
        return df['Email'].tolist()
    except Exception as e:
        mb.showerror("Error", e)
        ic(e)
        return []   
       
def read_body(body_path):
    try:
        with open(body_path, 'r', encoding='utf-8') as file:
            content = file.read()
            return content
    except FileNotFoundError:
        print(f"File not found: {body_path}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
# ----------------
def prepare_email_body(template_path, **replacements):
    
    try:
        with open(template_path, 'r', encoding='utf-8') as file:
            content = file.read()
            content = content.format_map(replacements)
            return content
    except FileNotFoundError:
        print(f"Error: The file '{template_path}' was not found.")
        return None
    except KeyError as e:
        print(f"Missing replacement for: {e}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    
# =================
# def main():

root = tk.Tk()

root.title("Data Processing [singh.jayram@in.ibm.com, pijavale@in.ibm.com ]")
root.option_add("*tearOff", False)
window_width = 1200
window_height = 800
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
center_x = int((screen_width / 2) - (window_width / 2))
center_y = int((screen_height / 2) - (window_height / 2))
root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
root.iconbitmap('J.ico')
# self.root.resizable(False, False)
style = ttk.Style(root)
root.tk.call('source', 'forest-dark.tcl')
style.theme_use('forest-dark')
supply_sheet, report_sheet, demand_sheet, utilization_sheet, skill_sheet, sender_name, notes = read_conf()
# ---------------
frame = ttk.Frame(root)
frame.pack()
# Entries with default values
# cp = os.path.join(os.getcwd(), 'ICO Bench_Roll Off_demands(19thJan24) .xlsx')

value1 = supply_sheet
value2 = report_sheet
value3 = demand_sheet
value4 = utilization_sheet
value5 = skill_sheet

style = ttk.Style()
style.configure("Report.TLabel", font=("Arial", 12), anchor="w")
style.configure("Report.Tvalue", font=("Arial", 12, "bold"), anchor="center")

# style.configure("Report.TLabel", font=("Arial", 12), anchor="w", background="gray", foreground="white")
# style.configure("Report.Tvalue", font=("Arial", 12, "bold"), anchor="center", background="gray")


# style.configure("main.TEntry", font=("Arial", 10), anchor="w")
style.configure("main.TLabel", font=("Arial", 10), anchor="w")
hyperlink_font = tkFont.Font(family='Helvetica', size=10, slant='italic', underline=True)

# style.map("C.TButton",
#     foreground=[('pressed', 'red'), ('active', 'blue')],
#     background=[('pressed', '!disabled', 'black'), ('active', 'white')]
#     )

# File information frame
head = ttk.Label(frame, text="Data Processing [v1.0.1]", font=("Arial", 18))
head.grid(row=0, column=0,columnspan=2, pady=20, sticky="NSEW")

# File process frame
p_frame = ttk.Labelframe(frame, text="File Process")
p_frame.grid(row=1, column=0, sticky="NSEW")

# File information frame
L1 = ttk.Label(p_frame, text="Supply sheet :")
L1.grid(row=0, column=0, padx=5, pady=5)
default_value1 = tk.StringVar()
default_value1.set(value1)
supply_sheet_ttk = ttk.Entry(p_frame, textvariable=default_value1, width=20)
supply_sheet_ttk.grid(row=0, column=1, padx=10, pady=5)
B1=ttk.Button(p_frame, text="Browse", command=browse_supply)
B1.grid(row=0, column=2, padx=5, pady=5)


# tower_list = get_tower_list(supply_file=supply_file)

t_label = ttk.Label(p_frame, text="Select Tower :")
t_label.grid(row=1, column=0, padx=10, pady=5)
selected_value = tk.StringVar()
selected_value.set(tower_list[0])
combobox_tower = ttk.Combobox(p_frame, textvariable=selected_value, values=tower_list, state="readonly", width=17)
combobox_tower.grid(row=1, column=1, padx=10, pady=5)
combobox_tower.bind("<<ComboboxSelected>>", lambda event: on_tower_select(selected_value.get()))

B1_combo = ttk.Button(p_frame, text="Get Tower", command=lambda: get_tower_list(supply_sheet=supply_sheet), width=10)
B1_combo.grid(row=1, column=2, padx=5, pady=5)

L2=ttk.Label(p_frame, text="Previous sheet :")
L2.grid(row=2, column=0, padx=10, pady=5)
default_value2 = tk.StringVar()
default_value2.set(value2)
report_sheet_ttk = ttk.Entry(p_frame, textvariable=default_value2, width=20)
report_sheet_ttk.grid(row=2, column=1, padx=10, pady=5)
B2=ttk.Button(p_frame, text="Browse", command=browse_tracker,width=10)
B2.grid(row=2, column=2, padx=5, pady=5)

L3=ttk.Label(p_frame, text="Demand sheet :")
L3.grid(row=3, column=0, padx=10, pady=5)
default_value3 = tk.StringVar()
default_value3.set(value3)
demand_sheet_ttk = ttk.Entry(p_frame, textvariable=default_value3, width=20)
demand_sheet_ttk.grid(row=3, column=1, padx=10, pady=5)
B3=ttk.Button(p_frame, text="Browse", command=browse_dm,width=10)
B3.grid(row=3, column=2, padx=5, pady=5)

L4=ttk.Label(p_frame, text="Utilization sheet :")
L4.grid(row=4, column=0, padx=10, pady=5)
default_value4 = tk.StringVar()
default_value4.set(value4)
utilization_sheet_ttk = ttk.Entry(p_frame, textvariable=default_value4, width=20)
utilization_sheet_ttk.grid(row=4, column=1, padx=10, pady=5)
B4=ttk.Button(p_frame, text="Browse", command=browse_util,width=10)
B4.grid(row=4, column=2, padx=5, pady=5)

L5=ttk.Label(p_frame, text="Skill sheet :")
L5.grid(row=5, column=0, padx=10, pady=5)
default_value5 = tk.StringVar()
default_value5.set(value5)
skill_sheet_ttk = ttk.Entry(p_frame, textvariable=default_value5, width=20)
skill_sheet_ttk.grid(row=5, column=1, padx=10, pady=5)
B4=ttk.Button(p_frame, text="Browse", command=browse_skill,width=10)
B4.grid(row=5, column=2, padx=5, pady=5)


# Process button
B5=ttk.Button(p_frame, text="Process", command=GetData)
B5.grid(row=6, column=0, padx=10, pady=10)

# self.hyperlink_enabled = True

hyperlink_label = tk.Label(p_frame, text=New_Report_file, fg="white", cursor="hand2", font=hyperlink_font)
hyperlink_label.grid(row=7, column=0, columnspan=2,  padx=10, pady=10)
hyperlink_label.bind("<Button-1>")
hyperlink_label = hyperlink_label
# self.hyperlink_enabled = False

B6=ttk.Button(p_frame, text="Open Dashboard", command=dashboard)
B6.grid(row=7, column=2, padx=5, pady=5)

# Report

report_frame = ttk.Labelframe(frame, text="Report Date: {}", width=600, height=200)
report_frame.grid(row=1, column=1,rowspan=3, padx=10, sticky="NSEW")

plt_frame1 = ttk.Labelframe(report_frame, width=100, height=100)
plt_frame1.grid(row=0, column=0, padx=10, sticky="NSEW")

plt_frame2 = ttk.Labelframe(report_frame, width=100, height=100)
plt_frame2.grid(row=1, column=0, padx=10, sticky="NSEW")

data_frame = ttk.Labelframe(report_frame, width=200, height=200)
data_frame.grid(row=2, column=0, padx=10, sticky="NSEW")

bench_count = ttk.Label(data_frame, text="Bench Count:", style="Report.TLabel")
upcoming_bench = ttk.Label(data_frame, text="Upcoming Bench:", style="Report.TLabel" )
total = ttk.Label(data_frame, text="Total:", style="Report.TLabel")
bench_count.grid(row=0, column=0)
upcoming_bench.grid(row=1, column=0)
total.grid(row=2, column=0)

bench_count_value = ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
upcoming_bench_count_value = ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
total_count = ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
bench_count_value.grid(row=0, column=1)
upcoming_bench_count_value.grid(row=1, column=1)
total_count.grid(row=2, column=1)

AVAILABLE_WITHIN_15_DAYS= ttk.Label(data_frame, text="Available within 15 days:", style="Report.TLabel")
AVAILABLE_IN_16_TO_30_DAYS= ttk.Label(data_frame, text="Available in 16 to 30 days:", style="Report.TLabel")
AVAILABLE_IN_31_TO_60_DAYS= ttk.Label(data_frame, text="Available in 31 to 60 days:", style="Report.TLabel")
AVAILABLE_IN_61_TO_90_DAYS= ttk.Label(data_frame, text="Available in 61 to 90_days:", style="Report.TLabel")
AVAILABLE_AFTER_90_DAYS= ttk.Label(data_frame, text="Aavailable after 90 days:", style="Report.TLabel")

AVAILABLE_WITHIN_15_DAYS.grid(row=3, column=0)
AVAILABLE_IN_16_TO_30_DAYS.grid(row=4, column=0)
AVAILABLE_IN_31_TO_60_DAYS.grid(row=5, column=0)
AVAILABLE_IN_61_TO_90_DAYS.grid(row=6, column=0)
AVAILABLE_AFTER_90_DAYS.grid(row=7, column=0)

vAVAILABLE_WITHIN_15_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel" , width=10)
vAVAILABLE_IN_16_TO_30_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
vAVAILABLE_IN_31_TO_60_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
vAVAILABLE_IN_61_TO_90_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
vAVAILABLE_AFTER_90_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)

vAVAILABLE_WITHIN_15_DAYS.grid(row=3, column=1)
vAVAILABLE_IN_16_TO_30_DAYS.grid(row=4, column=1)
vAVAILABLE_IN_31_TO_60_DAYS.grid(row=5, column=1)
vAVAILABLE_IN_61_TO_90_DAYS.grid(row=6, column=1)
vAVAILABLE_AFTER_90_DAYS.grid(row=7, column=1)

#  CONFIGURE
widgets = [bench_count, upcoming_bench, total, AVAILABLE_WITHIN_15_DAYS,AVAILABLE_IN_16_TO_30_DAYS ,AVAILABLE_IN_31_TO_60_DAYS,AVAILABLE_IN_61_TO_90_DAYS,AVAILABLE_AFTER_90_DAYS]
            # , vAVAILABLE_WITHIN_15_DAYS, vAVAILABLE_IN_31_TO_60_DAYS, vAVAILABLE_IN_61_TO_90_DAYS, vAVAILABLE_AFTER_90_DAYS]

padx, pady, width = 3, 3, 30
for i, widget in enumerate(widgets):
    widget.grid(row=i, column=0, padx=padx, pady=pady)
    widget.config(width=width)


# email frame
email_frame = ttk.Labelframe(frame, text="Email")
email_frame.grid(row=2, column=0, pady=10, sticky="NSEW")
L3 = ttk.Label(email_frame, text="Select PDL sheet:")
L3.grid(row=0, column=0, padx=10, pady=5)
PDL_sheet_name = tk.StringVar()
PDL_sheet_name.set(os.getcwd()+"\\recipient_list.xlsx")
E1 = ttk.Entry(email_frame, textvariable=PDL_sheet_name, font="Arial 10")
E1.grid(row=0, column=1, padx=10, pady=5)
B4=ttk.Button(email_frame, text="Browse", command=lambda: browse_PDL(E1),width=10)
B4.grid(row=0, column=2, padx=10, pady=5)
B5=ttk.Button(email_frame, text="Send email", command=SendEmailbutton)
B5.grid(row=1, column=0, padx=10, pady=10)

# Logs frame
output_frame = ttk.Labelframe(frame, text="Logs")
output_frame.grid(row=3, column=0,pady=5, sticky="NSEW")
scrolled_text = st.ScrolledText(output_frame, wrap=tk.WORD, width=60, height=5)
scrolled_text.grid(row=3,column=0)
# ---plt---


if len(New_Report_file) >= 1:
    plt1 = plt1_data
    plt2 = plt2_data    
else:
    plt1, plt2 = get_plt_data(key=plt_key)


pig1, ax1 = plt.subplots()
root_color = root.cget('bg')
title1 = ax1.set_title("Bench Data", color='white', fontsize=10)
ax1.pie(list(plt1.values()), labels=list(plt1.keys()), autopct='%1.1f%%', textprops={'fontsize': 8, 'color': 'white'} )
pig1.set_size_inches(3, 2)
pig1.set_facecolor(root_color)
title1.set_color('white')
title1.set_fontsize(10)

pig2, ax2 = plt.subplots()
title2 = ax2.set_title("Upcoming Bench Data", color='white', fontsize=10)
ax2.set_title("Upcoming Bench Data")
ax2.pie(list(plt2.values()), labels=list(plt2.keys()), autopct='%1.1f%%', textprops={'fontsize': 8, 'color': 'white'} )
pig2.set_size_inches(3, 2)
pig2.set_facecolor(root_color)
title2.set_color('white')
title2.set_fontsize(10)

canvas1 = FigureCanvasTkAgg(pig1, plt_frame1)
canvas1.draw()
canvas1.get_tk_widget().pack(side=tk.LEFT)

canvas2 = FigureCanvasTkAgg(pig2, plt_frame2)
canvas2.draw()
canvas2.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=False)

root.mainloop()

# if __name__ == "__main__":
#     main()
