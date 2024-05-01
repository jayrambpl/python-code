import openpyxl
import pandas as pd
from openpyxl import load_workbook, workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from openpyxl.utils import get_column_letter

import numpy as np
import datetime
pd.options.mode.chained_assignment = None

# -----------------------
def Excel_format(file_path, sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    blue_border = Border(
        left=Side(style='thin', color='0000FF'),
        right=Side(style='thin', color='0000FF'),
        top=Side(style='thin', color='0000FF'),
        bottom=Side(style='thin', color='0000FF')
    )

    align_left_center = Alignment(horizontal='left', vertical='center')

    for cell in sheet[1]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_left_center

    
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18  
        for cell in row:
            cell.border = blue_border
            cell.alignment = align_left_center

    for column in sheet.iter_cols():
        sheet.column_dimensions[column[0].column_letter].width = 15  

    wb.save(file_path)

# -----------------------
def Excel_process(supply_file, bench_file):
    
    df = pd.read_excel(supply_file)
    IoT_c = df['Tower'] == 'IoT'
    Bench_c = df['Roll Off Status'] == 'BENCH'
    IoT_bench_df = df[IoT_c & Bench_c]
    now = datetime.datetime.now()
    file_path = now.strftime("IoT_Bench_Data_%Y%m%d_%H%M%S.xlsx")
    
    IoT_bench_df_selected = IoT_bench_df.copy()

    columns_to_save = ['Serial', 'Practitioner Notes ID', 'JRSS-Primary', 'Tower', 'Roll Off Status', 'BandAsPerSR', 'PAL Dept', 'Reporting GDMIS - PAL Dept','Bench Aging' ] 
    
    IoT_bench_df_selected = IoT_bench_df[columns_to_save]
    
    new_columns = ['Lead', 'Action Owner', 'Status2', 'Proposed / Confirmed Cased', 'Remarks', 'Comparision']
    for col in new_columns:
        IoT_bench_df_selected.loc[:, col] = None
    
    IoT_bench_df_selected.insert(0, 'SN', range(1, 1 + len(IoT_bench_df_selected)))
    
    # Read Tracker 
    df_old_tracker = pd.read_excel(bench_file)
    
    df1_updated = IoT_bench_df_selected.drop(new_columns, axis=1, errors='ignore').merge(df_old_tracker[['Serial'] + new_columns], on='Serial', how='left')
  
    df1_updated.to_excel(file_path, sheet_name='IoT_Bench', index=False)

    bench_count = IoT_bench_df.__len__()
    output_str = f"\nNew Excel Report : {file_path} \n\nTotal Bench Count: {bench_count} "
    
    upcoming_bench = df['Roll Off Status'].str.contains('AVAILABLE')
    IoT_upcoming_bench = df[IoT_c & upcoming_bench]
    IoT_upcoming_bench_selected = IoT_upcoming_bench[columns_to_save]
    
    IoT_upcoming_bench_selected.insert(0, 'SN', range(1, 1 + len(IoT_upcoming_bench_selected)))
    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        IoT_upcoming_bench_selected.to_excel(writer, sheet_name='IoT_Upcoming_Bench', index=False)

    upcoming_count = IoT_upcoming_bench.__len__()
    output_str += f"\nUpcoming Bench Count: {upcoming_count}\n"
    
    Excel_format(file_path, "IoT_Bench")
    Excel_format(file_path, "IoT_Upcoming_Bench")

    return(output_str)

   #---------------- 
