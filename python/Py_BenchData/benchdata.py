import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.scrolledtext as st
import tkinter.messagebox as mb
import tkinter.font as tkFont

import win32com.client as client
import pandas as pd
import gc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import datetime
import subprocess
import py_email_template
import json
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
plt.rcParams["axes.prop_cycle"] = plt.cycler(color=["#00FF00", "#FF0000", "#0000FF", "#FFFF00", "#FF00FF"])

New_Report_file = ''
report_file = ''
report_date = datetime.datetime.now()
bench_count = 0
upcoming_bench_count = 0
available_within_15_days = 0
available_in_16_to_30_days = 0
available_in_31_to_60_days = 0
available_in_61_to_90_days = 0
available_after_90_days = 0
supply_file =''
Confirmed = 0
Proposed = 0 
Available = 0
sub_total = 0
plt_key = '2024-09-01 12:34:56'
# plt_key = ''
# ------------------------
def get_plt_data(key, bench_plt, upcoming_plt):
    
    db= read_history(key=key)

    bench_plt["CONFIRMED"] = db["CONFIRMED"]
    bench_plt["PROPOSED"] = db["PROPOSED"]
    bench_plt["AVAILABLE"] = db["AVAILABLE"]

    upcoming_plt["WITHIN15"] = db["WITHIN15"]  
    upcoming_plt["WITHIN30"] = db["WITHIN30"]
    upcoming_plt["WITHIN60"] = db["WITHIN60"]
    upcoming_plt["WITHIN90"] = db["WITHIN90"]
    upcoming_plt["AFTER90"] = db["AFTER90"]

    return bench_plt, upcoming_plt
# --------------------------
# -------------------------

# ----------------------

def open_excel_file(New_Report_file):
    file_to_open= os.getcwd() + "\\" +New_Report_file

    if os.path.exists(file_to_open):
        subprocess.Popen(["start", "excel", file_to_open], shell=True)
    else:
        mb.showinfo("Information", "Excel file not found.")

# ------------------------
def read_body(body_path):
    try:
        with open(body_path, 'r', encoding='utf-8') as file:
            content = file.read()
            return content
    except FileNotFoundError:
        mb.showerror("Error", "File not found.")
        return None
    except Exception as e:
        mb.showerror("Error", e)
        return None
# ------------------------
def save_history(key, json_data):
    history_path = os.path.join(os.getcwd(), "history.json")
    if not os.path.exists(history_path):
        mb.showerror("Error", "History database not found.")
        return None
    
    try:
        with open('history.json', 'r') as file:
            historical_data = json.load(file)
    except FileNotFoundError:
        historical_data = {}

    historical_data[key] = json_data
    try:
        with open('history.json', 'w') as file:
            json.dump(historical_data, file, indent=4)
    except Exception as e:
        mb.showerror("Error", e)
        return None 
    
def read_history(key):
    history_path = os.path.join(os.getcwd(), "history.json")
    if not os.path.exists(history_path):
        mb.showerror("Error", "History database not found.")
        return None
    
    try:
        with open('history.json', 'r') as file:
            historical_data = json.load(file)
    except FileNotFoundError:
        historical_data = {}

    if key in historical_data:
        return historical_data[key]
    else:
        return None
        
# ------------------------
def SendEmailbutton():
    global supply_file
    global bench_count
    global upcoming_bench_count
    global available_within_15_days
    global available_in_31_to_60_days
    global available_in_61_to_90_days
    global available_after_90_days

    if New_Report_file  == '':
        mb.showerror("Error", "No data has been processed today.\nSending a blank email may not be the best choice.")
        return None
    
    excel_path= E1.get()
    html_json_path =os.path.join(os.getcwd(), "body.htm")
    email_body = read_body(html_json_path)
    if not email_body:
        mb.showerror("Error", "Email body not found.")  
        return
    
    now = datetime.datetime.now()
    email_date = now.strftime("%Y-%d-%m %H:%M:%S")

    variables_list = { "REPORT_DATE": "09-01-2024 23:17",
                      "SUPPLY_FILE": "Jayram_data.xlsx",
                      "B_COUNT": "",
                      "CONFIRMED": "",
                      "PROPOSED": "",
                      "AVAILABLE": "",
                      "UP_COUNT": "",
                      "WITHIN15": "",
                      "WITHIN30": "",
                      "WITHIN60": "",
                      "WITHIN90": "",
                      "AFTER90": "",
                      "SUB_TOTAL": "",
                      "TOTAL": ""
    }

    variables_list["REPORT_DATE"] = email_date
    variables_list["SUPPLY_FILE"] = os.path.basename(supply_file)
    variables_list["B_COUNT"] = str(bench_count)
    variables_list["CONFIRMED"] = str(Confirmed)
    variables_list["PROPOSED"] = str(Proposed)
    variables_list["AVAILABLE"] = str(Available)
    variables_list["UP_COUNT"] = str(upcoming_bench_count)
    variables_list["WITHIN15"] = str(available_within_15_days)
    variables_list["WITHIN30"] = str(available_in_16_to_30_days)
    variables_list["WITHIN60"] = str(available_in_31_to_60_days)
    variables_list["WITHIN90"] = str(available_in_61_to_90_days)
    variables_list["AFTER90"] = str(available_after_90_days)
    variables_list["SUB_TOTAL"] = str(sub_total)
    variables_list["TOTAL"] = str(bench_count+upcoming_bench_count)

    for variable, value in variables_list.items():
        placeholder = f'%{variable}%'
        email_body = email_body.replace(placeholder, value)

    recipients = read_recipient(excel_path)
    attachment_path =os.path.join(os.getcwd(), New_Report_file)
    subject = f"IoT Bench data report for {email_date} with attachment for today's review."
    try:
        send_email(subject, email_body, attachment_path, recipients)
        mb.showinfo("Information", "Email Sent Successully.")
        scrolled_text.insert(tk.END, "Email Sent Successully.")
        scrolled_text.see(tk.END)
        scrolled_text.update()
        key = email_date
        save_history(key, variables_list )
    
    except Exception as e:
        mb.showerror("Error", e)
        return None

# ------------ --   
def browse_supply():

    file1 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select a file",
        filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*"))
    )
    if file1:
       supply_file_name.delete(0, tk.END)
       supply_file_name.insert(0, file1)

# ------------ --   
def browse_tracker():
    file2 = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select a file",
        filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*"))
    )
    if file2:
        bench_file_name.delete(0, tk.END)
        bench_file_name.insert(0, file2)            

# --------------
def browse_PDL(E1):
    pdl_file = filedialog.askopenfilename( 
        initialdir=os.getcwd(),
        title="Select a file",
        filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*"))
    )
    if pdl_file:
        E1.delete(0, tk.END)
        E1.insert(0, pdl_file)            

# ----------------
def GetData():
    global supply_file
    supply_file = supply_file_name.get()
    bench_file = bench_file_name.get()
    S_filename = os.path.basename(supply_file)
    B_filename = os.path.basename(bench_file)
    logs_str = f"Processing...\nSupply File: {S_filename} \nBench File: {B_filename}"
    scrolled_text.insert(tk.END, logs_str)
    
    # mb.showinfo("Information", supply_file + "\n " + bench_file)
    logstr= GetBenchData(supply_file, bench_file)
    # New_Report_file = BenchApp.GetBenchData(supply_file, bench_file,self)
    # report_manager.scrolled_text.insert(tk.END, logs_str)
    
    # self.hyperlink_enabled = True
    if os.path.exists(New_Report_file):
        # hyperlink_font = tkFont.Font(family='Helvetica', size=12, slant='italic', underline=True)
        hyperlink_label.config(text=New_Report_file)
        hyperlink_label.bind("<Button-1>", lambda event: open_excel_file(New_Report_file))
        hyperlink_label.config(fg="white") 
    else:
        hyperlink_label.config(text=New_Report_file)
        hyperlink_label.unbind("<Button-1>")
        hyperlink_label.config(fg="gray")  
# -----------------
def ExcelFormat(file_path, sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    header_fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    blue_border = Border(
        left=Side(style='thin', color='0000FF'),
        right=Side(style='thin', color='0000FF'),
        top=Side(style='thin', color='0000FF'),
        bottom=Side(style='thin', color='0000FF')
    )

    align_left_center = Alignment(horizontal='left', vertical='center')

    for cell in sheet[1]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_left_center
        if cell.column <= 12:
            cell.fill = header_fill_gray
        else:
            cell.fill = header_fill    
    
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18  
        for cell in row:
            cell.border = blue_border
            cell.alignment = align_left_center

    for column in sheet.iter_cols():
        # sheet.column_dimensions[column[0].column_letter].width = 15  
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)  
        sheet.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
# ------------------------
def GetBenchData(supply_file, bench_file):
    global New_Report_file
    global report_date
    global bench_count
    global upcoming_bench_count
    global available_within_15_days
    global available_in_16_to_30_days
    global available_in_31_to_60_days
    global available_in_61_to_90_days
    global available_after_90_days
    global Confirmed
    global Proposed
    global Available
    global sub_total


    pd.options.mode.chained_assignment = None
    now = datetime.datetime.now()
    report_date = now.strftime("%Y-%m-%d")
    New_Report_file = now.strftime("Report_%Y%m%d_%H%M%S.xlsx")

    New_sheet_bench = 'IoT_Bench'
    New_sheet_upcoming_bench = 'IoT_Upcoming_Bench'

    suppy_df = pd.read_excel(supply_file, sheet_name='Supply')

    # IoT_c = suppy_df['Skill'] == 'IoT'
    # Bench_c = suppy_df['Roll Off Status'] == 'BENCH'
    # IoT_bench_df = suppy_df[IoT_c & Bench_c]
        
    IoT_bench_df = suppy_df['Roll Off Status'] == 'BENCH'
    

    IoT_bench_df_selected = IoT_bench_df.copy()

    columns_to_save = ['Serial', 'Practitioner Notes ID', 'JRSS-Primary', 'Skill', 'Roll Off Status', 'BandAsPerSR', 'Bench Aging', 'PAL Dept', 'Reporting GDMIS - PAL Dept','Customer Name','Project Name'] 
    
    IoT_bench_df_selected = IoT_bench_df[columns_to_save]
    
    new_columns = ['Lead', 'Action Owner', 'Status2', 'Account Proposed','Proposed / Confirmed Cased', 'Remarks']
    for col in new_columns:
        IoT_bench_df_selected.loc[:, col] = None
    
    IoT_bench_df_selected.insert(0, 'SN', range(1, 1 + len(IoT_bench_df_selected)))
    
    # add remarks data from previous Tracker 
    old_tracker_df = pd.read_excel(bench_file, sheet_name= New_sheet_bench)
    
    new_mergered_df = IoT_bench_df_selected.drop(new_columns, axis=1, errors='ignore').merge(old_tracker_df[['Serial'] + new_columns], on='Serial', how='left')
    
    new_mergered_df.to_excel(New_Report_file, sheet_name= New_sheet_bench, index=False)

    bench_count = IoT_bench_df.__len__()
    logs_str = f"\nNew Excel Report : {New_Report_file} \n\nTotal Bench Count: {bench_count} "
    scrolled_text.insert(tk.END, logs_str)
    # Upcoming Bench # -------
    
    upcoming_bench_df = suppy_df['Roll Off Status'].str.contains('AVAILABLE')
    # IoT_upcoming_bench_df = suppy_df[IoT_c & upcoming_bench_df]

    IoT_upcoming_bench_selected = upcoming_bench_df[columns_to_save]
    new_columns = ['Lead', 'Action Owner', 'Status2', 'Proposed / Confirmed Cased', 'Remarks', 'Comparision']
    for col in new_columns:
        IoT_upcoming_bench_selected.loc[:, col] = None
    
    IoT_upcoming_bench_selected.insert(0, 'SN', range(1, 1 + len(IoT_upcoming_bench_selected)))

    # add data
    old_upcoming_data = pd.read_excel(bench_file, sheet_name= New_sheet_upcoming_bench)
    new_mergered_df = IoT_upcoming_bench_selected.drop(new_columns, axis=1, errors='ignore').merge(old_upcoming_data[['Serial'] + new_columns], on='Serial', how='left')

    with pd.ExcelWriter(New_Report_file, engine='openpyxl', mode='a') as writer:
        new_mergered_df.to_excel(writer, sheet_name=New_sheet_upcoming_bench, index=False)

    upcoming_bench_count = new_mergered_df.__len__()
    logs_str = f"\nUpcoming Bench Count: {upcoming_bench_count}\n"
    # app_instance.scrolled_text.insert(tk.END, logs_str)
    
    # fill report
    
    
    df_report = new_mergered_df.copy()
    
    mask_15_days = df_report['Roll Off Status'] == 'AVAILABLE WITHIN 15 DAYS'
    mast_16_30_days = df_report['Roll Off Status'] == 'AVAILABLE IN 16 TO 30 DAYS'
    mask_31_60_days = df_report['Roll Off Status'] == 'AVAILABLE IN 31 TO 60 DAYS'
    mask_61_90_days = df_report['Roll Off Status'] == 'AVAILABLE IN 61 TO 90 DAYS'
    mask_after_90_days = df_report['Roll Off Status'] == 'AVAILABLE AFTER 90 DAYS'

    # Applying the masks to the DataFrame and getting the counts
    available_within_15_days = df_report[mask_15_days].shape[0]
    available_in_16_to_30_days = df_report[mast_16_30_days].shape[0]
    available_in_31_to_60_days = df_report[mask_31_60_days].shape[0]
    available_in_61_to_90_days = df_report[mask_61_90_days].shape[0]
    available_after_90_days = df_report[mask_after_90_days].shape[0]

    sub_total = available_within_15_days + available_in_16_to_30_days + available_in_31_to_60_days + available_in_61_to_90_days

    str_dashboard = f"Report Date: {report_date}"
    report_frame.config(text=str_dashboard)
    bench_count_value.config(text=str(bench_count))
    upcoming_bench_count_value.config(text=str(upcoming_bench_count))
    total_count.config(text=str(bench_count+upcoming_bench_count))

    vAVAILABLE_WITHIN_15_DAYS.config(text=str(available_within_15_days))
    vAVAILABLE_IN_16_TO_30_DAYS.config(text=str(available_in_16_to_30_days))
    vAVAILABLE_IN_31_TO_60_DAYS.config(text=str(available_in_31_to_60_days))
    vAVAILABLE_IN_61_TO_90_DAYS.config(text=str(available_in_61_to_90_days))
    vAVAILABLE_AFTER_90_DAYS.config(text=str(available_after_90_days))
    logs_str = f"\nReport Generated\n-----------------\n"
    scrolled_text.insert(tk.END, logs_str)
    
    # ------------additional field------
    # kdf kdFK KHF
    #  HFWHFKAKKJK

    df_bench_sub = pd.read_excel(New_Report_file, sheet_name='IoT_Bench') 
    
    status2_filter = ['Confirmed', 'confirm']
    mask_df_confirmed = df_bench_sub[df_bench_sub['Status2'].isin(status2_filter) | df_bench_sub['Status2'].isna()]
    Confirmed = len(mask_df_confirmed)

    # mask_df_confirmed = df_bench_sub['Status2'].str.lower().isin(['confirmed', 'confirm', ''])
    # mask_df_confirmed = df_bench_sub['Status2'] == 'confirm'
    mask_df_proposed = df_bench_sub['Status2'] == 'Proposed'
    mask_df_available = df_bench_sub['Status2'] == 'Bench'

    # Confirmed = df_bench_sub[mask_df_confirmed].shape[0]
    Proposed = df_bench_sub[mask_df_proposed].shape[0]
    Available = df_bench_sub[mask_df_available].shape[0]

    # print(f"Confirmed: {Confirmed}, Proposed:{Proposed}, Available:{Available}")

    #  merge previoud data
    
    ExcelFormat(New_Report_file, "IoT_Bench")
    ExcelFormat(New_Report_file, "IoT_Upcoming_Bench")
    logs_str = f"\nExcel Formatted\n----------------\n"
    scrolled_text.insert(tk.END, logs_str)
    return(New_Report_file)  
# ------------------------
def send_email(subject, body, attachment_path, recipients):
    outlook = client.Dispatch('Outlook.Application')
    message = outlook.CreateItem(0) 
    message.Subject = subject
    
    message.HTMLBody = body
    
    if attachment_path and os.path.exists(attachment_path):
        message.Attachments.Add(Source=attachment_path)

    
    for recipient in recipients:
        message.Recipients.Add(recipient)

    message.Send()
    # print("Email sent to:", recipients)

def read_recipient(excel_path):
    df = pd.read_excel(excel_path)
    return df['Email'].tolist()

def read_body(body_path):
    try:
        with open(body_path, 'r', encoding='utf-8') as file:
            content = file.read()
            return content
    except FileNotFoundError:
        print(f"File not found: {body_path}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
# ----------------
def prepare_email_body(template_path, **replacements):
    
    try:
        with open(template_path, 'r', encoding='utf-8') as file:
            content = file.read()
            content = content.format_map(replacements)
            return content
    except FileNotFoundError:
        print(f"Error: The file '{template_path}' was not found.")
        return None
    except KeyError as e:
        print(f"Missing replacement for: {e}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    
# =================

root = tk.Tk()

root.title("Bench Data Processing")
root.option_add("*tearOff", False)
window_width = 1200
window_height = 800
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
center_x = int((screen_width / 2) - (window_width / 2))
center_y = int((screen_height / 2) - (window_height / 2))
root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
root.iconbitmap('J.ico')
# self.root.resizable(False, False)
style = ttk.Style(root)
root.tk.call('source', 'forest-dark.tcl')
style.theme_use('forest-dark')

# ---------------
frame = ttk.Frame(root)
frame.pack()
# Entries with default values
value1 = os.getcwd() + "\\Sheet_Supply.xlsx"
value2 = os.getcwd() + "\\IoT_Bench_Data_20240104_132537.xlsx"

style = ttk.Style()
style.configure("Report.TLabel", font=("Arial", 12), anchor="w")
style.configure("Report.Tvalue", font=("Arial", 12, "bold"), anchor="center")

# style.configure("Report.TLabel", font=("Arial", 12), anchor="w", background="gray", foreground="white")
# style.configure("Report.Tvalue", font=("Arial", 12, "bold"), anchor="center", background="gray")


# style.configure("main.TEntry", font=("Arial", 10), anchor="w")
style.configure("main.TLabel", font=("Arial", 10), anchor="w")
hyperlink_font = tkFont.Font(family='Helvetica', size=10, slant='italic', underline=True)

# style.map("C.TButton",
#     foreground=[('pressed', 'red'), ('active', 'blue')],
#     background=[('pressed', '!disabled', 'black'), ('active', 'white')]
#     )

# File information frame
head = ttk.Label(frame, text="Bench Data Processing", font=("Arial", 18))
head.grid(row=0, column=0,columnspan=2, pady=20, sticky="NSEW")

# File process frame
p_frame = ttk.Labelframe(frame, text="File Process")
p_frame.grid(row=1, column=0, sticky="NSEW")

# File information frame
L1 = ttk.Label(p_frame, text="Supply sheet name:")
L1.grid(row=0, column=0, padx=5, pady=5)
default_value1 = tk.StringVar()
default_value1.set(value1)
supply_file_name = ttk.Entry(p_frame, textvariable=default_value1, width=20)
supply_file_name.grid(row=0, column=1, padx=10, pady=5)
B1=ttk.Button(p_frame, text="Browse", command=browse_supply)
B1.grid(row=0, column=2, padx=5, pady=5)


L2=ttk.Label(p_frame, text="Previout sheet name:")
L2.grid(row=1, column=0, padx=10, pady=5)
default_value2 = tk.StringVar()
default_value2.set(value2)
bench_file_name = ttk.Entry(p_frame, textvariable=default_value2, width=20)
bench_file_name.grid(row=1, column=1, padx=10, pady=5)
B2=ttk.Button(p_frame, text="Browse", command=browse_tracker,width=10)
B2.grid(row=1, column=2, padx=5, pady=5)

# Process button
B3=ttk.Button(p_frame, text="Process", command=GetData)
B3.grid(row=2, column=0, padx=10, pady=10)

# Report

report_frame = ttk.Labelframe(frame, text="Report Date: {}", width=600, height=200)
report_frame.grid(row=1, column=1,rowspan=3, padx=10, sticky="NSEW")

plt_frame1 = ttk.Labelframe(report_frame, width=100, height=100)
plt_frame1.grid(row=0, column=0, padx=10, sticky="NSEW")

plt_frame2 = ttk.Labelframe(report_frame, width=100, height=100)
plt_frame2.grid(row=1, column=0, padx=10, sticky="NSEW")

data_frame = ttk.Labelframe(report_frame, width=200, height=200)
data_frame.grid(row=2, column=0, padx=10, sticky="NSEW")

bench_count = ttk.Label(data_frame, text="Bench Count:", style="Report.TLabel")
upcoming_bench = ttk.Label(data_frame, text="Upcoming Bench:", style="Report.TLabel" )
total = ttk.Label(data_frame, text="Total:", style="Report.TLabel")
bench_count.grid(row=0, column=0)
upcoming_bench.grid(row=1, column=0)
total.grid(row=2, column=0)

bench_count_value = ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
upcoming_bench_count_value = ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
total_count = ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
bench_count_value.grid(row=0, column=1)
upcoming_bench_count_value.grid(row=1, column=1)
total_count.grid(row=2, column=1)

AVAILABLE_WITHIN_15_DAYS= ttk.Label(data_frame, text="Available within 15 days:", style="Report.TLabel")
AVAILABLE_IN_16_TO_30_DAYS= ttk.Label(data_frame, text="Available in 16 to 30 days:", style="Report.TLabel")
AVAILABLE_IN_31_TO_60_DAYS= ttk.Label(data_frame, text="Available in 31 to 60 days:", style="Report.TLabel")
AVAILABLE_IN_61_TO_90_DAYS= ttk.Label(data_frame, text="Available in 61 to 90_days:", style="Report.TLabel")
AVAILABLE_AFTER_90_DAYS= ttk.Label(data_frame, text="Aavailable after 90 days:", style="Report.TLabel")

AVAILABLE_WITHIN_15_DAYS.grid(row=3, column=0)
AVAILABLE_IN_16_TO_30_DAYS.grid(row=4, column=0)
AVAILABLE_IN_31_TO_60_DAYS.grid(row=5, column=0)
AVAILABLE_IN_61_TO_90_DAYS.grid(row=6, column=0)
AVAILABLE_AFTER_90_DAYS.grid(row=7, column=0)

vAVAILABLE_WITHIN_15_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel" , width=10)
vAVAILABLE_IN_16_TO_30_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
vAVAILABLE_IN_31_TO_60_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
vAVAILABLE_IN_61_TO_90_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)
vAVAILABLE_AFTER_90_DAYS= ttk.Label(data_frame, text="0", style="Report.TLabel", width=10)

vAVAILABLE_WITHIN_15_DAYS.grid(row=3, column=1)
vAVAILABLE_IN_16_TO_30_DAYS.grid(row=4, column=1)
vAVAILABLE_IN_31_TO_60_DAYS.grid(row=5, column=1)
vAVAILABLE_IN_61_TO_90_DAYS.grid(row=6, column=1)
vAVAILABLE_AFTER_90_DAYS.grid(row=7, column=1)

#  CONFIGURE
widgets = [bench_count, upcoming_bench, total, AVAILABLE_WITHIN_15_DAYS,AVAILABLE_IN_16_TO_30_DAYS ,AVAILABLE_IN_31_TO_60_DAYS,AVAILABLE_IN_61_TO_90_DAYS,AVAILABLE_AFTER_90_DAYS]
            # , vAVAILABLE_WITHIN_15_DAYS, vAVAILABLE_IN_31_TO_60_DAYS, vAVAILABLE_IN_61_TO_90_DAYS, vAVAILABLE_AFTER_90_DAYS]

padx, pady, width = 3, 3, 30
for i, widget in enumerate(widgets):
    widget.grid(row=i, column=0, padx=padx, pady=pady)
    widget.config(width=width)
   

# self.hyperlink_enabled = True

hyperlink_label = tk.Label(p_frame, text=New_Report_file, fg="white", cursor="hand2", font=hyperlink_font)
hyperlink_label.grid(row=3, column=0, columnspan=3,  padx=10, pady=10)
hyperlink_label.bind("<Button-1>")
hyperlink_label = hyperlink_label
# self.hyperlink_enabled = False


# email frame
email_frame = ttk.Labelframe(frame, text="Email")
email_frame.grid(row=2, column=0, pady=10, sticky="NSEW")
L3 = ttk.Label(email_frame, text="Select PDL sheet:")
L3.grid(row=0, column=0, padx=10, pady=5)
PDL_sheet_name = tk.StringVar()
PDL_sheet_name.set(os.getcwd()+"\\recipient_list.xlsx")
E1 = ttk.Entry(email_frame, textvariable=PDL_sheet_name, font="Arial 10")
E1.grid(row=0, column=1, padx=10, pady=5)
B4=ttk.Button(email_frame, text="Browse", command=lambda: browse_PDL(E1),width=10)
B4.grid(row=0, column=2, padx=10, pady=5)
B5=ttk.Button(email_frame, text="Send email", command=SendEmailbutton)
B5.grid(row=1, column=0, padx=10, pady=10)


# Logs frame
output_frame = ttk.Labelframe(frame, text="Logs")
output_frame.grid(row=3, column=0,pady=5, sticky="NSEW")
scrolled_text = st.ScrolledText(output_frame, wrap=tk.WORD, width=60, height=5)
scrolled_text.grid(row=3,column=0)
# ---plt---

bench = { 
                "CONFIRMED": "",
                "PROPOSED": "",
                "AVAILABLE": "",
}
upcoming = {
                "WITHIN15": "",
                "WITHIN30": "",
                "WITHIN60": "",
                "WITHIN90": "",
                "AFTER90": ""
}


if plt_key:
    plt1, plt2 = get_plt_data(key=plt_key,bench_plt=bench,upcoming_plt=upcoming)
    pig1, ax1 = plt.subplots()
    
    root_color = root.cget('bg')
    title1 = ax1.set_title("Bench Data", color='white', fontsize=10)
    ax1.pie(list(plt1.values()), labels=list(plt1.keys()), autopct='%1.1f%%', textprops={'fontsize': 8, 'color': 'white'} )
    pig1.set_size_inches(3, 2)
    pig1.set_facecolor(root_color)
    title1.set_color('white')
    title1.set_fontsize(10)

    pig2, ax2 = plt.subplots()
    title2 = ax2.set_title("Upcoming Bench Data", color='white', fontsize=10)
    ax2.set_title("Upcoming Bench Data")
    ax2.pie(list(plt2.values()), labels=list(plt2.keys()), autopct='%1.1f%%', textprops={'fontsize': 8, 'color': 'white'} )
    pig2.set_size_inches(3, 2)
    pig2.set_facecolor(root_color)
    title2.set_color('white')
    title2.set_fontsize(10)

    canvas1 = FigureCanvasTkAgg(pig1, plt_frame1)
    canvas1.draw()
    canvas1.get_tk_widget().pack(side=tk.LEFT)

    canvas2 = FigureCanvasTkAgg(pig2, plt_frame2)
    canvas2.draw()
    canvas2.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=False)
root.mainloop()
